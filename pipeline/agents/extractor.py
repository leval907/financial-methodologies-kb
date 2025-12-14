#!/usr/bin/env python3
"""
Agent A: Document Extractor
Извлекает и нормализует текст из различных форматов документов

Поддерживаемые форматы:
- PDF (текстовый и отсканированный)
- DOCX
- PPTX
- XLSX
- TXT, MD
"""

import sys
import subprocess
from pathlib import Path
from typing import Dict, Any, Optional, List
import json
from datetime import datetime

import openpyxl
import pandas as pd


class DocumentExtractor:
    """Извлекает текст из различных типов документов"""
    
    SUPPORTED_FORMATS = {
        '.pdf': 'extract_pdf',
        '.docx': 'extract_docx',
        '.pptx': 'extract_pptx',
        '.xlsx': 'extract_xlsx',
        '.xls': 'extract_xlsx',
        '.txt': 'extract_text',
        '.md': 'extract_text',
    }
    
    def __init__(self, use_markitdown: bool = True, max_excel_rows: int = 1000):
        """
        Args:
            use_markitdown: Использовать markitdown для PDF/DOCX/PPTX
            max_excel_rows: Максимум строк для Excel таблиц
        """
        self.use_markitdown = use_markitdown
        self.max_excel_rows = max_excel_rows
    
    def extract(self, filepath: Path) -> Dict[str, Any]:
        """
        Извлечь содержимое документа
        
        Returns:
            {
                'content': str,  # Основной текст в Markdown
                'metadata': dict,  # Метаданные
                'tables': list,  # Таблицы (для Excel)
                'formulas': list,  # Формулы (для Excel)
            }
        """
        if not filepath.exists():
            raise FileNotFoundError(f"File not found: {filepath}")
        
        suffix = filepath.suffix.lower()
        if suffix not in self.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {suffix}")
        
        method_name = self.SUPPORTED_FORMATS[suffix]
        method = getattr(self, method_name)
        
        result = method(filepath)
        result['metadata']['extracted_at'] = datetime.now().isoformat()
        result['metadata']['source_file'] = str(filepath)
        result['metadata']['format'] = suffix
        
        return result
    
    def extract_pdf(self, filepath: Path) -> Dict[str, Any]:
        """Извлечь текст из PDF"""
        if self.use_markitdown:
            try:
                return self._extract_via_markitdown(filepath)
            except RuntimeError as e:
                # Если markitdown не справился, используем native extraction
                print(f"⚠️  markitdown failed: {e}, falling back to PyMuPDF")
        
        # Native extraction через PyMuPDF
        return self._extract_pdf_native(filepath)
    
    def extract_docx(self, filepath: Path) -> Dict[str, Any]:
        """
        Извлечь текст из DOCX с fallback стратегией:
        1. markitdown (быстро)
        2. mammoth (только текст, надёжно)
        """
        # Стратегия 1: markitdown
        if self.use_markitdown:
            try:
                return self._extract_via_markitdown(filepath)
            except Exception as e:
                print(f"⚠️ markitdown failed for DOCX: {e}, trying fallback...")
        
        # Стратегия 2: mammoth (fallback)
        try:
            import mammoth
            
            with open(filepath, 'rb') as f:
                result = mammoth.extract_raw_text(f)
            
            return {
                'content': result.value,
                'metadata': {
                    'method': 'mammoth',
                    'lines': len(result.value.splitlines()),
                    'chars': len(result.value),
                    'quality': 'basic',
                    'note': 'Formatting lost, text only (fallback method)'
                },
                'tables': [],
                'formulas': [],
            }
        except Exception as e:
            raise RuntimeError(f"All DOCX extraction methods failed: {e}")
    
    def extract_pptx(self, filepath: Path) -> Dict[str, Any]:
        """Извлечь текст из PPTX"""
        if self.use_markitdown:
            return self._extract_via_markitdown(filepath)
        else:
            raise NotImplementedError("Native PPTX extraction not implemented yet")
    
    def extract_xlsx(self, filepath: Path) -> Dict[str, Any]:
        """Извлечь данные из Excel"""
        wb = openpyxl.load_workbook(filepath, data_only=False)
        
        content_parts = [f"# {filepath.name}\n\n"]
        tables = []
        formulas = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Метаданные листа
            content_parts.append(f"## {sheet_name}\n\n")
            content_parts.append(f"- Dimensions: `{sheet.dimensions}`\n")
            content_parts.append(f"- Rows: {sheet.max_row}, Columns: {sheet.max_column}\n\n")
            
            # Данные в Markdown таблицу
            try:
                data = list(sheet.values)
                if data:
                    cols = data[0]
                    df = pd.DataFrame(data[1:], columns=cols)
                    
                    # Ограничить строки
                    if len(df) > self.max_excel_rows:
                        df_preview = df.head(self.max_excel_rows)
                        content_parts.append(
                            f"*Showing first {self.max_excel_rows} of {len(df)} rows*\n\n"
                        )
                    else:
                        df_preview = df
                    
                    content_parts.append(df_preview.to_markdown(index=False))
                    content_parts.append("\n\n")
                    
                    # Сохранить таблицу для отдельного анализа
                    tables.append({
                        'sheet': sheet_name,
                        'dimensions': sheet.dimensions,
                        'rows': sheet.max_row,
                        'columns': sheet.max_column,
                        'data': df.to_dict('records')[:100],  # Первые 100 строк
                    })
            except Exception as e:
                content_parts.append(f"*Error extracting sheet data: {e}*\n\n")
            
            # Извлечь формулы
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # formula
                        formulas.append({
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'formula': cell.value,
                            'result': cell.internal_value,
                        })
        
        return {
            'content': ''.join(content_parts),
            'metadata': {
                'sheets': wb.sheetnames,
                'sheet_count': len(wb.sheetnames),
                'total_formulas': len(formulas),
            },
            'tables': tables,
            'formulas': formulas[:100],  # Первые 100 формул
        }
    
    def extract_text(self, filepath: Path) -> Dict[str, Any]:
        """Извлечь текст из TXT/MD"""
        content = filepath.read_text(encoding='utf-8')
        
        return {
            'content': content,
            'metadata': {
                'encoding': 'utf-8',
                'lines': len(content.splitlines()),
                'chars': len(content),
            },
            'tables': [],
            'formulas': [],
        }
    
    def _extract_via_markitdown(self, filepath: Path) -> Dict[str, Any]:
        """Использовать markitdown CLI для извлечения"""
        try:
            result = subprocess.run(
                ['markitdown', str(filepath)],
                capture_output=True,
                text=True,
                timeout=60,  # 1 минута таймаут
            )
            
            if result.returncode != 0:
                raise RuntimeError(f"markitdown failed: {result.stderr}")
            
            content = result.stdout
            
            # Проверяем качество вывода
            if len(content.strip()) < 50:
                raise RuntimeError("markitdown output too short, possibly failed")
            
            return {
                'content': content,
                'metadata': {
                    'method': 'markitdown',
                    'lines': len(content.splitlines()),
                    'chars': len(content),
                    'quality': 'good',
                },
                'tables': [],
                'formulas': [],
            }
        except subprocess.TimeoutExpired:
            raise RuntimeError(f"markitdown timeout for {filepath}")
        except FileNotFoundError:
            raise RuntimeError("markitdown not installed")
    
    def _extract_pdf_native(self, filepath: Path) -> Dict[str, Any]:
        """Извлечь текст из PDF через PyMuPDF (fitz)"""
        try:
            import fitz
        except ImportError:
            raise RuntimeError("PyMuPDF not installed. Run: pip install PyMuPDF")
        
        doc = fitz.open(filepath)
        
        content_parts = [f"# {filepath.name}\n\n"]
        total_chars = 0
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            page_text = page.get_text()
            
            if page_text.strip():
                content_parts.append(f"## Page {page_num + 1}\n\n")
                content_parts.append(page_text)
                content_parts.append("\n\n---\n\n")
                total_chars += len(page_text)
        
        doc.close()
        
        content = ''.join(content_parts)
        
        # Проверяем что получили текст
        if total_chars < 100:
            raise RuntimeError(f"PDF appears to be scanned (no text layer). Only {total_chars} chars extracted.")
        
        return {
            'content': content,
            'metadata': {
                'method': 'PyMuPDF',
                'pages': len(doc),
                'lines': len(content.splitlines()),
                'chars': len(content),
                'quality': 'good',
            },
            'tables': [],
            'formulas': [],
        }


def process_document(
    input_path: Path,
    output_dir: Path,
    book_id: str,
    use_markitdown: bool = True
) -> Dict[str, Any]:
    """
    Обработать документ и сохранить результаты
    
    Args:
        input_path: Путь к исходному документу
        output_dir: Директория для сохранения (sources/<book_id>/)
        book_id: Идентификатор книги
        use_markitdown: Использовать markitdown
    
    Returns:
        Словарь с путями к созданным файлам
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    
    extractor = DocumentExtractor(use_markitdown=use_markitdown)
    result = extractor.extract(input_path)
    
    # Сохранить основной текст
    text_file = output_dir / 'raw_text.md'
    text_file.write_text(result['content'], encoding='utf-8')
    
    # Сохранить метаданные
    metadata_file = output_dir / 'metadata.json'
    metadata = {
        'book_id': book_id,
        'source_file': str(input_path),
        'format': input_path.suffix,
        **result['metadata']
    }
    metadata_file.write_text(json.dumps(metadata, indent=2, ensure_ascii=False))
    
    # Сохранить таблицы (если есть)
    if result['tables']:
        tables_dir = output_dir / 'tables'
        tables_dir.mkdir(exist_ok=True)
        
        for i, table in enumerate(result['tables']):
            table_file = tables_dir / f"table_{i+1}_{table['sheet']}.json"
            table_file.write_text(json.dumps(table, indent=2, ensure_ascii=False))
    
    # Сохранить формулы (если есть)
    if result['formulas']:
        formulas_file = output_dir / 'formulas.json'
        formulas_file.write_text(
            json.dumps(result['formulas'], indent=2, ensure_ascii=False)
        )
    
    return {
        'text_file': str(text_file),
        'metadata_file': str(metadata_file),
        'tables_count': len(result['tables']),
        'formulas_count': len(result['formulas']),
    }


def main():
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Agent A: Extract text from documents'
    )
    parser.add_argument('input', type=Path, help='Input file')
    parser.add_argument('--output-dir', type=Path, required=True,
                       help='Output directory (sources/<book_id>/)')
    parser.add_argument('--book-id', required=True,
                       help='Book identifier')
    parser.add_argument('--no-markitdown', action='store_true',
                       help='Do not use markitdown')
    
    args = parser.parse_args()
    
    try:
        result = process_document(
            args.input,
            args.output_dir,
            args.book_id,
            use_markitdown=not args.no_markitdown
        )
        
        print(f"✅ Extracted: {args.book_id}")
        print(f"   Text: {result['text_file']}")
        print(f"   Metadata: {result['metadata_file']}")
        if result['tables_count']:
            print(f"   Tables: {result['tables_count']}")
        if result['formulas_count']:
            print(f"   Formulas: {result['formulas_count']}")
    
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
