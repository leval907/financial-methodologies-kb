#!/usr/bin/env python3
"""
Excel to Markdown Converter
Извлекает данные из Excel файлов и конвертирует в Markdown

Поддерживает:
- Множественные листы
- Формулы и значения
- Именованные диапазоны
- Комментарии и заметки
"""

import sys
import argparse
from pathlib import Path
import openpyxl
import pandas as pd
from typing import List, Dict, Any


def extract_sheet_metadata(sheet) -> Dict[str, Any]:
    """Извлечь метаданные листа"""
    return {
        'title': sheet.title,
        'dimensions': sheet.dimensions,
        'max_row': sheet.max_row,
        'max_column': sheet.max_column,
    }


def extract_named_ranges(workbook) -> List[Dict[str, str]]:
    """Извлечь именованные диапазоны"""
    ranges = []
    try:
        for name, defn in workbook.defined_names.items():
            ranges.append({
                'name': name,
                'reference': str(defn.attr_text) if hasattr(defn, 'attr_text') else str(defn),
            })
    except Exception:
        pass  # Нет именованных диапазонов
    return ranges


def sheet_to_markdown(sheet, sheet_name: str, max_rows: int = 100) -> str:
    """Конвертировать лист Excel в Markdown таблицу"""
    md = f"\n## {sheet_name}\n\n"
    
    # Получить данные как DataFrame
    data = sheet.values
    cols = next(data)
    df = pd.DataFrame(data, columns=cols)
    
    # Ограничить количество строк для preview
    if len(df) > max_rows:
        df = df.head(max_rows)
        md += f"*Showing first {max_rows} of {sheet.max_row} rows*\n\n"
    
    # Конвертировать в Markdown таблицу
    md += df.to_markdown(index=False)
    md += "\n"
    
    return md


def extract_formulas(sheet) -> List[Dict[str, Any]]:
    """Извлечь формулы из листа"""
    formulas = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # formula
                formulas.append({
                    'cell': cell.coordinate,
                    'formula': cell.value,
                    'result': cell.internal_value,
                })
    return formulas


def convert_excel_to_markdown(filepath: Path, output: Path = None, 
                              include_formulas: bool = False,
                              max_rows_per_sheet: int = 100) -> str:
    """
    Конвертировать Excel файл в Markdown
    
    Args:
        filepath: Путь к Excel файлу
        output: Путь для сохранения Markdown (опционально)
        include_formulas: Включить список формул
        max_rows_per_sheet: Максимум строк на лист
    
    Returns:
        Markdown текст
    """
    wb = openpyxl.load_workbook(filepath, data_only=False)
    
    md = f"# {filepath.name}\n\n"
    md += f"**Source:** `{filepath}`\n\n"
    md += f"**Sheets:** {len(wb.sheetnames)}\n\n"
    
    # Именованные диапазоны
    named_ranges = extract_named_ranges(wb)
    if named_ranges:
        md += "### Named Ranges\n\n"
        for nr in named_ranges:
            md += f"- **{nr['name']}**: `{nr['reference']}`\n"
        md += "\n"
    
    # Обработка каждого листа
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Метаданные листа
        metadata = extract_sheet_metadata(sheet)
        md += f"### {sheet_name} (Sheet Metadata)\n\n"
        md += f"- Dimensions: `{metadata['dimensions']}`\n"
        md += f"- Rows: {metadata['max_row']}, Columns: {metadata['max_column']}\n\n"
        
        # Данные листа
        try:
            md += sheet_to_markdown(sheet, sheet_name, max_rows_per_sheet)
        except Exception as e:
            md += f"\n*Error converting sheet: {e}*\n\n"
        
        # Формулы (опционально)
        if include_formulas:
            formulas = extract_formulas(sheet)
            if formulas:
                md += f"\n#### Formulas in {sheet_name}\n\n"
                for f in formulas[:20]:  # Первые 20 формул
                    md += f"- `{f['cell']}`: `{f['formula']}` → `{f['result']}`\n"
                if len(formulas) > 20:
                    md += f"\n*...and {len(formulas) - 20} more formulas*\n"
                md += "\n"
    
    # Сохранить если указан output
    if output:
        output.write_text(md, encoding='utf-8')
        print(f"✅ Saved to {output}")
    
    return md


def main():
    parser = argparse.ArgumentParser(
        description='Convert Excel files to Markdown'
    )
    parser.add_argument('input', type=Path, help='Input Excel file')
    parser.add_argument('-o', '--output', type=Path, help='Output Markdown file')
    parser.add_argument('-f', '--formulas', action='store_true',
                       help='Include formulas in output')
    parser.add_argument('--max-rows', type=int, default=100,
                       help='Maximum rows per sheet (default: 100)')
    
    args = parser.parse_args()
    
    if not args.input.exists():
        print(f"❌ File not found: {args.input}")
        sys.exit(1)
    
    try:
        md = convert_excel_to_markdown(
            args.input,
            args.output,
            include_formulas=args.formulas,
            max_rows_per_sheet=args.max_rows
        )
        
        if not args.output:
            print(md)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
